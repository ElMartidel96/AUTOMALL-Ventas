#!/usr/bin/env python3
"""
AUTOMALL — Generador de Control de Ventas (Excel)
Genera un archivo Excel profesional para el seguimiento de ventas de carros.

Ejecutar: python3 scripts/generate-sales-excel.py
Salida:   AUTOMALL-Control-Ventas.xlsx
"""

import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, Rule
    from openpyxl.styles.differential import DifferentialStyle
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, Rule
    from openpyxl.styles.differential import DifferentialStyle

# ================================================================
# PALETA DE COLORES AUTOMALL
# ================================================================
C_BLUE      = "1B3A6B"   # am-blue      — headers principales
C_BLUE_L    = "2B5EA7"   # am-blue-light — sub-headers
C_ORANGE    = "E8832A"   # am-orange    — acento / CTA
C_ORANGE_L  = "F5A623"   # am-orange-light
C_GREEN     = "2D8F4E"   # am-green     — éxito / pagado
C_DARK      = "0D1B2A"   # am-dark      — fondo oscuro
C_WHITE     = "FFFFFF"
C_LGRAY     = "F4F6FA"   # filas alternas claras
C_MGRAY     = "DDE2EC"   # bordes
C_DGRAY     = "5A6A7E"   # texto secundario
C_RED       = "C62828"   # repo / vencido
C_RED_L     = "FFCDD2"   # fondo rojo claro
C_YELLOW    = "F9A825"   # pendiente / warning
C_YELLOW_L  = "FFF9C4"   # fondo amarillo claro
C_GREEN_L   = "C8E6C9"   # fondo verde claro (pagado)
C_BLUE_LL   = "E8EEF7"   # fondo azul muy claro

# ================================================================
# HELPERS DE ESTILO
# ================================================================
def mk_fill(color):
    return PatternFill(fill_type="solid", fgColor=color)

def mk_font(color=C_DARK, size=10, bold=False, italic=False):
    return Font(name="Calibri", color=color, size=size, bold=bold, italic=italic)

def mk_border(color=C_MGRAY, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(cell, text, bg=C_BLUE, fg=C_WHITE, size=10, bold=True, wrap=True):
    cell.value = text
    cell.fill = mk_fill(bg)
    cell.font = mk_font(fg, size, bold)
    cell.alignment = mk_align("center", "center", wrap)
    cell.border = mk_border(bg)

def style_data(cell, bg=C_WHITE, fg=C_DARK, size=10, bold=False,
               h="center", fmt=None):
    cell.fill = mk_fill(bg)
    cell.font = mk_font(fg, size, bold)
    cell.alignment = mk_align(h, "center")
    cell.border = mk_border()
    if fmt:
        cell.number_format = fmt

# Formatos numéricos
FMT_DATE  = "MM/DD/YYYY"
FMT_MONEY = '"$"#,##0.00'
FMT_NUM   = "#,##0"

# ================================================================
# CREAR WORKBOOK
# ================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ================================================================
# HOJA 1 — REGISTRO DE VENTAS
# ================================================================
ws = wb.create_sheet("Ventas")
ws.sheet_properties.tabColor = C_BLUE
ws.sheet_view.showGridLines = False

# ------- Definición de columnas -------
# (letra, encabezado, ancho, categoría)
COLS = [
    # IDENTIFICACIÓN
    ("A",  "#",                      4,  "id"),
    ("B",  "FECHA\nVENTA",          13,  "info"),
    # CLIENTE
    ("C",  "NOMBRE\nCLIENTE",       26,  "client"),
    ("D",  "TELÉFONO",              15,  "client"),
    ("E",  "EMAIL",                 24,  "client"),
    ("F",  "REFERIDO POR",          20,  "client"),
    ("G",  "FUENTE",                13,  "client"),
    # VEHÍCULO
    ("H",  "MARCA",                 13,  "vehicle"),
    ("I",  "MODELO",                13,  "vehicle"),
    ("J",  "AÑO",                    7,  "vehicle"),
    ("K",  "COLOR",                 11,  "vehicle"),
    ("L",  "VIN",                   20,  "vehicle"),
    ("M",  "MILLAJE",               11,  "vehicle"),
    # FINANCIERO
    ("N",  "PRECIO\nVENTA",         13,  "finance"),
    ("O",  "ENGANCHE",              13,  "finance"),
    ("P",  "MONTO\nFINANCIADO",     14,  "finance"),
    # PICK PAYMENT
    ("Q",  "PICK\nPAYMENT",          9,  "pick"),
    ("R",  "# CUOTAS",              10,  "pick"),
    ("S",  "1er PAGO\n(FECHA)",     13,  "pick"),
    ("T",  "DÍA DE\nPAGO",         13,  "pick"),
    ("U",  "CUOTA\nSEMANAL",       13,  "pick"),
    # PAGOS 1–6 (fecha + estatus)
    ("V",  "PAGO 1\nFECHA",        13,  "pagos"),
    ("W",  "P1\nESTATUS",          11,  "pagos"),
    ("X",  "PAGO 2\nFECHA",        13,  "pagos"),
    ("Y",  "P2\nESTATUS",          11,  "pagos"),
    ("Z",  "PAGO 3\nFECHA",        13,  "pagos"),
    ("AA", "P3\nESTATUS",          11,  "pagos"),
    ("AB", "PAGO 4\nFECHA",        13,  "pagos"),
    ("AC", "P4\nESTATUS",          11,  "pagos"),
    ("AD", "PAGO 5\nFECHA",        13,  "pagos"),
    ("AE", "P5\nESTATUS",          11,  "pagos"),
    ("AF", "PAGO 6\nFECHA",        13,  "pagos"),
    ("AG", "P6\nESTATUS",          11,  "pagos"),
    # RESUMEN FINANCIERO
    ("AH", "TOTAL\nCOBRADO",       14,  "totals"),
    ("AI", "SALDO\nPENDIENTE",     14,  "totals"),
    # ADICIONAL
    ("AJ", "FINANCIERA /\nBANCO",  18,  "extra"),
    ("AK", "GPS\nINSTAL.",          9,  "extra"),
    ("AL", "ESTATUS\nDEAL",        13,  "extra"),
    ("AM", "COMISIÓN",             13,  "extra"),
    ("AN", "NOTAS",                35,  "extra"),
]

# Colores por categoría (bg, fg)
CAT_COLOR = {
    "id":      (C_DARK,   C_WHITE),
    "info":    (C_BLUE,   C_WHITE),
    "client":  (C_BLUE,   C_WHITE),
    "vehicle": (C_BLUE_L, C_WHITE),
    "finance": ("1A5C2B", C_WHITE),   # verde oscuro
    "pick":    (C_ORANGE, C_WHITE),
    "pagos":   ("4A148C", C_WHITE),   # morado
    "totals":  (C_GREEN,  C_WHITE),
    "extra":   (C_DGRAY,  C_WHITE),
}

# ------- Anchos de columna -------
for col_l, _, width, _ in COLS:
    ws.column_dimensions[col_l].width = width

# ------- FILA 1: Título principal -------
ws.row_dimensions[1].height = 42
ws.merge_cells("A1:AN1")
tc = ws["A1"]
tc.value = "AUTOMALL  ·  CONTROL DE VENTAS"
tc.fill = mk_fill(C_DARK)
tc.font = Font(name="Calibri", color=C_ORANGE, size=20, bold=True)
tc.alignment = mk_align("center", "center")

# ------- FILA 2: Grupos de categorías -------
ws.row_dimensions[2].height = 16
cat_groups = [
    ("A2:A2",   "ID",               "id"),
    ("B2:B2",   "INFO",             "info"),
    ("C2:G2",   "CLIENTE",          "client"),
    ("H2:M2",   "VEHÍCULO",         "vehicle"),
    ("N2:P2",   "FINANCIERO",       "finance"),
    ("Q2:U2",   "PICK PAYMENT",     "pick"),
    ("V2:AG2",  "CALENDARIO DE PAGOS", "pagos"),
    ("AH2:AI2", "RESUMEN",          "totals"),
    ("AJ2:AN2", "ADICIONAL",        "extra"),
]
for rng, label, cat in cat_groups:
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    bg, fg = CAT_COLOR[cat]
    cell.value = label
    cell.fill = mk_fill(bg)
    cell.font = mk_font(fg, 8, True)
    cell.alignment = mk_align("center", "center")

# ------- FILA 3: Encabezados de columna -------
ws.row_dimensions[3].height = 50
for col_l, header, _, cat in COLS:
    cell = ws[f"{col_l}3"]
    bg, fg = CAT_COLOR[cat]
    style_header(cell, header, bg=bg, fg=fg, size=9, bold=True, wrap=True)

# Congelar filas 1-3 y columna A-B
ws.freeze_panes = "C4"

# ------- FILAS DE DATOS: 4 a 153 (150 clientes) -------
STATUS_COLS = ["W", "Y", "AA", "AC", "AE", "AG"]

for r in range(4, 154):
    is_even = (r % 2 == 0)
    row_bg  = C_LGRAY if is_even else C_WHITE

    ws.row_dimensions[r].height = 20

    # ── Fórmulas automáticas ──────────────────────────────────

    # A: Número de fila (solo si hay cliente)
    f_a = ws[f"A{r}"]
    f_a.value = f'=IF(C{r}<>"",ROW()-3,"")'

    # P: Monto Financiado = Precio - Enganche (solo si Pick Payment = SÍ)
    f_p = ws[f"P{r}"]
    f_p.value = f'=IF(AND(Q{r}="SÍ",N{r}<>"",O{r}<>""),N{r}-O{r},"")'
    f_p.number_format = FMT_MONEY

    # T: Día de pago (calculado desde la fecha del 1er pago, locale-independent)
    f_t = ws[f"T{r}"]
    f_t.value = (
        f'=IF(S{r}<>"",'
        f'CHOOSE(WEEKDAY(S{r},2),'
        f'"Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"),'
        f'"")'
    )

    # U: Cuota semanal = Monto Financiado / # Cuotas
    f_u = ws[f"U{r}"]
    f_u.value = f'=IF(AND(P{r}<>"",R{r}>0),ROUND(P{r}/R{r},2),"")'
    f_u.number_format = FMT_MONEY

    # V: Pago 1 Fecha = 1er Pago Fecha (si Pick Payment = SÍ)
    f_v = ws[f"V{r}"]
    f_v.value = f'=IF(AND(Q{r}="SÍ",S{r}<>""),S{r},"")'
    f_v.number_format = FMT_DATE

    # X→AF: Fechas de pago 2→6 (cada +7 días desde el anterior)
    pay_pairs = [("X", "V"), ("Z", "X"), ("AB", "Z"), ("AD", "AB"), ("AF", "AD")]
    for cur_col, prev_col in pay_pairs:
        fc = ws[f"{cur_col}{r}"]
        fc.value = f'=IF({prev_col}{r}<>"",{prev_col}{r}+7,"")'
        fc.number_format = FMT_DATE

    # AH: Total Cobrado = cantidad de pagos marcados PAGADO × cuota
    ws[f"AH{r}"].value = (
        f'=IF(U{r}<>"",'
        f'(COUNTIF(W{r},"PAGADO")+COUNTIF(Y{r},"PAGADO")+COUNTIF(AA{r},"PAGADO")+'
        f'COUNTIF(AC{r},"PAGADO")+COUNTIF(AE{r},"PAGADO")+COUNTIF(AG{r},"PAGADO"))'
        f'*U{r},"")'
    )
    ws[f"AH{r}"].number_format = FMT_MONEY

    # AI: Saldo Pendiente = Monto Financiado - Total Cobrado
    ws[f"AI{r}"].value = (
        f'=IF(AND(P{r}<>"",AH{r}<>""),P{r}-AH{r},'
        f'IF(P{r}<>"",P{r},""))'
    )
    ws[f"AI{r}"].number_format = FMT_MONEY

    # ── Estilo base para todas las celdas de la fila ──────────
    for col_l, _, _, cat in COLS:
        cell = ws[f"{col_l}{r}"]
        # No sobreescribir celdas que ya tienen fórmula asignada
        bg, fg = CAT_COLOR[cat]
        cell.fill = mk_fill(row_bg)
        cell.font = mk_font(C_DARK, 10)
        cell.border = mk_border()
        cell.alignment = mk_align("center", "center")

    # Formatos específicos de celda
    ws[f"B{r}"].number_format = FMT_DATE
    ws[f"M{r}"].number_format = FMT_NUM
    ws[f"N{r}"].number_format = FMT_MONEY
    ws[f"O{r}"].number_format = FMT_MONEY
    ws[f"AM{r}"].number_format = FMT_MONEY

    # Notas alineadas a la izquierda
    ws[f"AN{r}"].alignment = mk_align("left", "center", wrap=True)

    # Color de la columna A (número de fila)
    ws[f"A{r}"].fill = mk_fill(C_BLUE_L if is_even else C_BLUE)
    ws[f"A{r}"].font = mk_font(C_WHITE, 9, bold=True)
    ws[f"A{r}"].alignment = mk_align("center", "center")

# ── Fila de TOTALES al final ──────────────────────────────────
LAST_DATA = 153
TOT_ROW   = 155
ws.row_dimensions[TOT_ROW].height = 22

ws.merge_cells(f"A{TOT_ROW}:M{TOT_ROW}")
tc = ws[f"A{TOT_ROW}"]
tc.value = "TOTALES"
tc.fill = mk_fill(C_BLUE)
tc.font = mk_font(C_WHITE, 10, bold=True)
tc.alignment = mk_align("center", "center")
tc.border = mk_border(C_BLUE)

totals_def = [
    ("N", f"=SUM(N4:N{LAST_DATA})", FMT_MONEY, "Precio total vendido"),
    ("O", f"=SUM(O4:O{LAST_DATA})", FMT_MONEY, "Total enganches"),
    ("P", f"=SUM(P4:P{LAST_DATA})", FMT_MONEY, "Total financiado"),
    ("AH", f"=SUM(AH4:AH{LAST_DATA})", FMT_MONEY, "Total cobrado"),
    ("AI", f"=SUM(AI4:AI{LAST_DATA})", FMT_MONEY, "Total saldo pendiente"),
    ("AM", f"=SUM(AM4:AM{LAST_DATA})", FMT_MONEY, "Total comisiones"),
]
for col_l, formula, fmt, _ in totals_def:
    cell = ws[f"{col_l}{TOT_ROW}"]
    cell.value = formula
    cell.number_format = fmt
    cell.fill = mk_fill(C_BLUE)
    cell.font = mk_font(C_WHITE, 10, bold=True)
    cell.alignment = mk_align("center", "center")
    cell.border = mk_border(C_BLUE)

# ================================================================
# VALIDACIONES DE DATOS (listas desplegables)
# ================================================================

def add_dv(ws, col_range, formula, title="", msg=""):
    dv = DataValidation(
        type="list",
        formula1=formula,
        showDropDown=False,
        showInputMessage=bool(msg),
        promptTitle=title,
        prompt=msg,
        showErrorMessage=True,
        errorTitle="Valor no válido",
        error="Selecciona un valor de la lista.",
    )
    ws.add_data_validation(dv)
    dv.add(col_range)
    return dv

add_dv(ws, "Q4:Q153",  '"SÍ,NO"',
        "Pick Payment", "¿El cliente paga en cuotas semanales?")

add_dv(ws, "G4:G153",
        '"REFERIDO,WALK-IN,INTERNET,REPEAT,REDES SOCIALES,ANUNCIO,OTRO"',
        "Fuente", "¿Cómo llegó el cliente?")

add_dv(ws, "AK4:AK153", '"SÍ,NO"',
        "GPS", "¿Se instaló GPS al vehículo?")

add_dv(ws, "AL4:AL153",
        '"ACTIVO,PAGADO COMPLETO,EN REPO,CANCELADO,RECOMPRADO"',
        "Estatus Deal", "Estado actual del deal")

# Estatus de pagos (P1 a P6)
for s_col in STATUS_COLS:
    add_dv(ws, f"{s_col}4:{s_col}153",
           '"PAGADO,VENCIDO,NO PAGÓ,PENDIENTE"',
           "Estatus Pago", "Marca el estado de este pago")

# Número de cuotas (1-12)
dv_cuotas = DataValidation(
    type="whole",
    operator="between",
    formula1="1",
    formula2="12",
    showErrorMessage=True,
    errorTitle="Cuotas inválidas",
    error="Ingresa un número entre 1 y 12.",
)
ws.add_data_validation(dv_cuotas)
dv_cuotas.add("R4:R153")

# ================================================================
# FORMATO CONDICIONAL
# ================================================================

# Helper: regla de color para estatus
def cf_status(ws, rng, value, bg_color, fg_color):
    dxf = DifferentialStyle(
        fill=mk_fill(bg_color),
        font=Font(name="Calibri", color=fg_color, bold=True, size=10)
    )
    rule = Rule(type="cellIs", operator="equal", formula=[f'"{value}"'], dxf=dxf)
    ws.conditional_formatting.add(rng, rule)

# Colores de estatus de pago
for s_col in STATUS_COLS:
    rng = f"{s_col}4:{s_col}153"
    cf_status(ws, rng, "PAGADO",   C_GREEN_L,  "1B5E20")
    cf_status(ws, rng, "VENCIDO",  C_RED_L,    C_RED)
    cf_status(ws, rng, "NO PAGÓ",  "EF9A9A",   "7F0000")
    cf_status(ws, rng, "PENDIENTE", C_YELLOW_L, "E65100")

# Colores de estatus del deal
cf_status(ws, "AL4:AL153", "ACTIVO",         "BBDEFB", "0D47A1")
cf_status(ws, "AL4:AL153", "PAGADO COMPLETO", C_GREEN_L, "1B5E20")
cf_status(ws, "AL4:AL153", "EN REPO",        C_RED_L,   C_RED)
cf_status(ws, "AL4:AL153", "CANCELADO",      "F5F5F5",  C_DGRAY)

# Saldo = 0 → verde (fully paid)
dxf_zero = DifferentialStyle(
    fill=mk_fill(C_GREEN_L),
    font=Font(name="Calibri", color="1B5E20", bold=True, size=10)
)
ws.conditional_formatting.add(
    "AI4:AI153",
    Rule(type="cellIs", operator="equal", formula=["0"], dxf=dxf_zero)
)

# Saldo > 0 → azul claro
dxf_pos = DifferentialStyle(fill=mk_fill(C_BLUE_LL))
ws.conditional_formatting.add(
    "AI4:AI153",
    Rule(type="cellIs", operator="greaterThan", formula=["0"], dxf=dxf_pos)
)

# Pick Payment = SÍ → resaltar celda Q en naranja
dxf_pick = DifferentialStyle(
    fill=mk_fill("FFF3E0"),
    font=Font(name="Calibri", color=C_ORANGE, bold=True, size=10)
)
ws.conditional_formatting.add(
    "Q4:Q153",
    Rule(type="cellIs", operator="equal", formula=['"SÍ"'], dxf=dxf_pick)
)

# ================================================================
# HOJA 2 — DASHBOARD
# ================================================================
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_properties.tabColor = C_ORANGE
ws2.sheet_view.showGridLines = False

for col, w in [("A",3),("B",28),("C",20),("D",20),("E",20),("F",20)]:
    ws2.column_dimensions[col].width = w

# Título
ws2.row_dimensions[1].height = 44
ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "AUTOMALL  ·  DASHBOARD DE VENTAS"
t.fill = mk_fill(C_DARK)
t.font = Font(name="Calibri", color=C_ORANGE, size=18, bold=True)
t.alignment = mk_align("center", "center")

# Subtítulo
ws2.row_dimensions[2].height = 18
ws2.merge_cells("A2:F2")
s = ws2["A2"]
s.value = "Los números se actualizan automáticamente desde la hoja 'Ventas'"
s.fill = mk_fill(C_BLUE)
s.font = Font(name="Calibri", color=C_WHITE, size=9, italic=True)
s.alignment = mk_align("center", "center")

# ── KPIs principales (fila 4-5) ──────────────────────────────
ws2.row_dimensions[3].height = 10
ws2.row_dimensions[4].height = 28
ws2.row_dimensions[5].height = 42

kpis = [
    ("B", "TOTAL VENTAS",
     f"=COUNTA(Ventas!C4:C{LAST_DATA})", None),
    ("C", "PICK PAYMENT",
     f'=COUNTIF(Ventas!Q4:Q{LAST_DATA},"SÍ")', None),
    ("D", "VENTAS ACTIVAS",
     f'=COUNTIF(Ventas!AL4:AL{LAST_DATA},"ACTIVO")', None),
    ("E", "TOTAL VENDIDO (PRECIO)",
     f"=SUM(Ventas!N4:N{LAST_DATA})", FMT_MONEY),
    ("F", "TOTAL COBRADO",
     f"=SUM(Ventas!AH4:AH{LAST_DATA})", FMT_MONEY),
]
for col, label, formula, fmt in kpis:
    hc = ws2[f"{col}4"]
    hc.value = label
    hc.fill = mk_fill(C_BLUE)
    hc.font = mk_font(C_WHITE, 9, True)
    hc.alignment = mk_align("center", "center", wrap=True)
    hc.border = mk_border(C_BLUE)

    vc = ws2[f"{col}5"]
    vc.value = formula
    vc.fill = mk_fill(C_LGRAY)
    vc.font = Font(name="Calibri", color=C_BLUE, size=20, bold=True)
    vc.alignment = mk_align("center", "center")
    vc.border = mk_border()
    if fmt:
        vc.number_format = fmt

# ── Métricas secundarias ──────────────────────────────────────
ws2.row_dimensions[7].height = 22
ws2.merge_cells("B7:C7")
hdr = ws2["B7"]
hdr.value = "MÉTRICA"
hdr.fill = mk_fill(C_BLUE)
hdr.font = mk_font(C_WHITE, 9, True)
hdr.alignment = mk_align("center", "center")
hdr.border = mk_border(C_BLUE)

ws2.merge_cells("D7:F7")
hdr2 = ws2["D7"]
hdr2.value = "VALOR"
hdr2.fill = mk_fill(C_BLUE)
hdr2.font = mk_font(C_WHITE, 9, True)
hdr2.alignment = mk_align("center", "center")
hdr2.border = mk_border(C_BLUE)

metrics = [
    (f'=COUNTIF(Ventas!AL4:AL{LAST_DATA},"PAGADO COMPLETO")',
     None,     "Deals PAGADOS COMPLETO",    C_GREEN_L,  "1B5E20"),
    (f'=COUNTIF(Ventas!AL4:AL{LAST_DATA},"EN REPO")',
     None,     "Deals EN REPO",             C_RED_L,    C_RED),
    (f'=COUNTIF(Ventas!AL4:AL{LAST_DATA},"CANCELADO")',
     None,     "Deals CANCELADOS",          "F5F5F5",   C_DGRAY),
    (f"=SUM(Ventas!AI4:AI{LAST_DATA})",
     FMT_MONEY,"Saldo Total PENDIENTE",     C_YELLOW_L, "E65100"),
    (f"=SUM(Ventas!AM4:AM{LAST_DATA})",
     FMT_MONEY,"Total COMISIONES",          C_GREEN_L,  "1B5E20"),
    (f'=COUNTIF(Ventas!W4:AG{LAST_DATA},"VENCIDO")',
     None,     "Pagos VENCIDOS (total)",    C_RED_L,    C_RED),
    (f'=COUNTIF(Ventas!W4:AG{LAST_DATA},"PENDIENTE")',
     None,     "Pagos PENDIENTES",          C_YELLOW_L, "E65100"),
    (f"=IFERROR(AVERAGE(Ventas!N4:N{LAST_DATA}),0)",
     FMT_MONEY,"Precio Promedio por Venta", C_BLUE_LL,  C_BLUE),
    (f"=IFERROR(AVERAGE(Ventas!O4:O{LAST_DATA}),0)",
     FMT_MONEY,"Enganche Promedio",         C_BLUE_LL,  C_BLUE),
]
for i, (formula, fmt, label, bg, fg) in enumerate(metrics):
    row = 8 + i
    ws2.row_dimensions[row].height = 22

    ws2.merge_cells(f"B{row}:C{row}")
    lc = ws2[f"B{row}"]
    lc.value = label
    lc.fill = mk_fill(C_BLUE if i % 2 == 0 else C_BLUE_L)
    lc.font = mk_font(C_WHITE, 10, bold=True)
    lc.alignment = mk_align("left", "center")
    lc.border = mk_border()

    ws2.merge_cells(f"D{row}:F{row}")
    vc = ws2[f"D{row}"]
    vc.value = formula
    vc.fill = mk_fill(bg)
    vc.font = Font(name="Calibri", color=fg, size=12, bold=True)
    vc.alignment = mk_align("center", "center")
    vc.border = mk_border()
    if fmt:
        vc.number_format = fmt

# ================================================================
# HOJA 3 — INSTRUCCIONES
# ================================================================
ws3 = wb.create_sheet("Instrucciones")
ws3.sheet_properties.tabColor = C_GREEN
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 60
ws3.column_dimensions["D"].width = 5

# Título
ws3.row_dimensions[1].height = 44
ws3.merge_cells("A1:D1")
t = ws3["A1"]
t.value = "AUTOMALL  ·  GUÍA DE USO"
t.fill = mk_fill(C_DARK)
t.font = Font(name="Calibri", color=C_GREEN, size=18, bold=True)
t.alignment = mk_align("center", "center")

instructions = [
    ("COLUMNAS AUTOMÁTICAS", None,
     "Estas columnas se calculan solas — NO las edites manualmente:"),
    ("Monto Financiado (P)", "#",
     "= Precio Venta − Enganche  (solo si Pick Payment = SÍ)"),
    ("Día de Pago (T)", "#",
     "= Día de semana del 1er pago (Lunes, Martes…) — calculado automático"),
    ("Cuota Semanal (U)", "#",
     "= Monto Financiado ÷ # Cuotas"),
    ("Fechas Pago 2–6 (X,Z,AB,AD,AF)", "#",
     "= Fecha anterior + 7 días (mismo día de la semana siempre)"),
    ("Total Cobrado (AH)", "#",
     "= Suma de cuotas marcadas como PAGADO"),
    ("Saldo Pendiente (AI)", "#",
     "= Monto Financiado − Total Cobrado"),
    ("", None, ""),
    ("CÓMO REGISTRAR UNA VENTA", None, "Sigue estos pasos en orden:"),
    ("Paso 1 — Fecha", "#", "Ingresa la fecha de la venta en columna B (formato MM/DD/AAAA)"),
    ("Paso 2 — Cliente", "#", "Llena nombre, teléfono, email, referido por, y fuente (dropdown)"),
    ("Paso 3 — Vehículo", "#", "Marca, modelo, año, color, VIN, y millaje"),
    ("Paso 4 — Precio y Enganche", "#", "Ingresa precio de venta y enganche en columnas N y O"),
    ("Paso 5 — Pick Payment", "#",
     "Selecciona SÍ o NO en columna Q. Si es SÍ, ingresa # cuotas (default 6) y la fecha del 1er pago"),
    ("Paso 6 — Pagos semanales", "#",
     "Las fechas de pago se calculan automáticamente. Solo cambia el ESTATUS cuando el cliente pague: "
     "PAGADO / VENCIDO / NO PAGÓ / PENDIENTE"),
    ("Paso 7 — Cierre", "#",
     "Cuando se complete, cambia Estatus Deal a PAGADO COMPLETO"),
    ("", None, ""),
    ("COLORES DE REFERENCIA", None, "Significado de colores en columnas de estatus:"),
    ("Verde claro", "#", "PAGADO — El cliente realizó el pago correctamente"),
    ("Rojo claro", "#", "VENCIDO — La fecha de pago pasó y no pagó"),
    ("Rojo oscuro", "#", "NO PAGÓ — Se confirmó que no va a pagar"),
    ("Amarillo", "#", "PENDIENTE — Pago próximo o en proceso"),
    ("", None, ""),
    ("CONSEJOS", None, "Mejores prácticas para mantener el control:"),
    ("GPS", "#", "Registra si instalaste GPS en columna AK — clave para recuperaciones"),
    ("Notas", "#", "Usa la columna AN para observaciones: carros trade-in, garantía, situación especial"),
    ("Comisión", "#", "Registra tu comisión en columna AM para llevar el control de ingresos"),
    ("Dashboard", "#", "Revisa la hoja 'Dashboard' para ver resumen general de todas las ventas"),
    ("Backup", "#", "Guarda una copia del archivo al final de cada semana"),
]

row_i = 3
for field, marker, desc in instructions:
    ws3.row_dimensions[row_i].height = 20 if field else 8

    if marker is None and field:  # Sección
        hc = ws3[f"B{row_i}"]
        hc.value = f"  {field}"
        hc.fill = mk_fill(C_BLUE)
        hc.font = mk_font(C_WHITE, 10, bold=True)
        hc.alignment = mk_align("left", "center")
        hc.border = mk_border(C_BLUE)

        dc = ws3[f"C{row_i}"]
        dc.value = f"  {desc}"
        dc.fill = mk_fill(C_BLUE_L)
        dc.font = mk_font(C_WHITE, 9, italic=True)
        dc.alignment = mk_align("left", "center")
        dc.border = mk_border(C_BLUE_L)

    elif field:  # Fila normal
        bg = C_LGRAY if row_i % 2 == 0 else C_WHITE

        fc = ws3[f"B{row_i}"]
        fc.value = field
        fc.fill = mk_fill(bg)
        fc.font = mk_font(C_BLUE_L, 10, bold=True)
        fc.alignment = mk_align("left", "center")
        fc.border = mk_border()

        dc = ws3[f"C{row_i}"]
        dc.value = desc
        dc.fill = mk_fill(bg)
        dc.font = mk_font(C_DARK, 10)
        dc.alignment = mk_align("left", "center", wrap=True)
        dc.border = mk_border()

    row_i += 1

# ================================================================
# HOJA 4 — CATÁLOGOS
# ================================================================
ws4 = wb.create_sheet("Catálogos")
ws4.sheet_properties.tabColor = C_DGRAY
ws4.sheet_view.showGridLines = False

for col, w in [("A",3),("B",22),("C",22),("D",22),("E",22)]:
    ws4.column_dimensions[col].width = w

ws4.row_dimensions[1].height = 38
ws4.merge_cells("A1:E1")
ct = ws4["A1"]
ct.value = "CATÁLOGOS — Listas de referencia para el Excel"
ct.fill = mk_fill(C_DARK)
ct.font = Font(name="Calibri", color=C_WHITE, size=14, bold=True)
ct.alignment = mk_align("center", "center")

catalogs = [
    ("B", "MARCAS DE AUTOS", [
        "Toyota", "Ford", "Chevrolet", "Nissan", "GMC",
        "Dodge", "Jeep", "Mazda", "Mercedes-Benz", "Honda",
        "Hyundai", "Kia", "BMW", "Volkswagen", "Chrysler",
        "RAM", "Cadillac", "Buick", "Lincoln", "Acura",
        "Subaru", "Mitsubishi", "Lexus", "Audi", "Infiniti",
    ]),
    ("C", "FUENTES DE VENTA", [
        "REFERIDO", "WALK-IN", "INTERNET", "REPEAT",
        "REDES SOCIALES", "ANUNCIO", "CRAIGSLIST",
        "FACEBOOK MARKETPLACE", "OTRO",
    ]),
    ("D", "FINANCIERAS / BANCOS", [
        "Cash (Contado)", "DriveTime", "CarMax Auto Finance",
        "Westlake Financial", "Ally Financial", "Capital One Auto",
        "CUDL", "CAC (Credit Acceptance)", "JM Family",
        "Santander Consumer", "TD Auto Finance",
        "Exeter Finance", "AmeriCredit (GM Financial)",
        "Southeast Toyota Finance", "Otro",
    ]),
    ("E", "REFERIDOS FRECUENTES", [
        "— Agregar nombres aquí —",
        "(Ejemplo: Juan Pérez)",
        "(Ejemplo: María García)",
        "(Ejemplo: Carlos López)",
    ]),
]

for col, header, items in catalogs:
    ws4.row_dimensions[3].height = 26
    hc = ws4[f"{col}3"]
    hc.value = header
    hc.fill = mk_fill(C_BLUE)
    hc.font = mk_font(C_WHITE, 10, bold=True)
    hc.alignment = mk_align("center", "center")
    hc.border = mk_border(C_BLUE)

    for i, item in enumerate(items):
        row = 4 + i
        ws4.row_dimensions[row].height = 18
        rc = ws4[f"{col}{row}"]
        rc.value = item
        rc.fill = mk_fill(C_LGRAY if i % 2 == 0 else C_WHITE)
        rc.font = mk_font(C_DARK, 10)
        rc.alignment = mk_align("left", "center")
        rc.border = mk_border()

# ================================================================
# FILA DE EJEMPLO EN HOJA VENTAS
# ================================================================
ws["B4"] = "01/15/2026"
ws["B4"].number_format = FMT_DATE
ws["C4"] = "María García (EJEMPLO)"
ws["D4"] = "(713) 555-0100"
ws["E4"] = "mgarcia@email.com"
ws["F4"] = "Juan López"
ws["G4"] = "REFERIDO"
ws["H4"] = "Toyota"
ws["I4"] = "Camry"
ws["J4"] = 2022
ws["K4"] = "Blanco"
ws["L4"] = "1HGBH41JXMN109186"
ws["M4"] = 45000
ws["M4"].number_format = FMT_NUM
ws["N4"] = 18500
ws["N4"].number_format = FMT_MONEY
ws["O4"] = 2000
ws["O4"].number_format = FMT_MONEY
ws["Q4"] = "SÍ"
ws["R4"] = 6
ws["S4"] = "01/22/2026"
ws["S4"].number_format = FMT_DATE
ws["W4"] = "PAGADO"
ws["Y4"] = "PAGADO"
ws["AA4"] = "PENDIENTE"
ws["AJ4"] = "Westlake Financial"
ws["AK4"] = "SÍ"
ws["AL4"] = "ACTIVO"
ws["AM4"] = 750
ws["AM4"].number_format = FMT_MONEY
ws["AN4"] = "Cliente de referido de Juan López. Trade-in: Honda Civic 2018."

# Estilo para la fila de ejemplo (fondo diferente para distinguirla)
for col_l, _, _, _ in COLS:
    cell = ws[f"{col_l}4"]
    cell.fill = mk_fill("FFF8E1")   # amarillo muy suave = fila de ejemplo

# ================================================================
# GUARDAR ARCHIVO
# ================================================================
out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
output_path = os.path.join(out_dir, "AUTOMALL-Control-Ventas.xlsx")

wb.save(output_path)

print("=" * 60)
print("✅  AUTOMALL — Control de Ventas generado exitosamente")
print("=" * 60)
print(f"📁  Archivo: {output_path}")
print(f"📋  Hoja 1 'Ventas'       — 150 filas, {len(COLS)} columnas")
print(f"📊  Hoja 2 'Dashboard'    — KPIs y métricas automáticas")
print(f"📖  Hoja 3 'Instrucciones'— Guía de uso paso a paso")
print(f"📂  Hoja 4 'Catálogos'    — Marcas, fuentes, financieras")
print()
print("🔧  Características incluidas:")
print("    • Fechas de pago 1-6 calculadas automáticamente (+7 días cada una)")
print("    • Día de semana del pago calculado desde fecha del 1er pago")
print("    • Monto por cuota = Monto financiado ÷ # cuotas")
print("    • Total cobrado y saldo pendiente en tiempo real")
print("    • Dropdowns: Pick Payment, Estatus pago, Fuente, GPS, Deal")
print("    • Formato condicional: verde=pagado, rojo=vencido, amarillo=pendiente")
print("    • Fila de ejemplo pre-cargada (fila 4 en amarillo claro)")
print("    • Fila de totales al final de la lista")
print()
print("🔑  Columnas automáticas (NO editar):")
print("    P = Monto Financiado | T = Día de Pago | U = Cuota Semanal")
print("    V,X,Z,AB,AD,AF = Fechas Pago 1-6")
print("    AH = Total Cobrado | AI = Saldo Pendiente")
